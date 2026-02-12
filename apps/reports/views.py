from django.shortcuts import render
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.db.models import Sum, F, Count
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from apps.inventory.models import Inventory
from apps.movements.models import Movement, Kardex
from apps.products.models import Product
from apps.warehouses.models import Warehouse
from datetime import datetime

@login_required
@permission_required('reports.view_report', raise_exception=True)
def dashboard(request):
    """Dashboard principal con estadísticas"""
    # Estadísticas para el dashboard
    total_products = Product.objects.filter(company=request.user.company, is_deleted=False).count()
    total_warehouses = Warehouse.objects.filter(company=request.user.company, is_deleted=False).count()
    low_stock_count = Inventory.objects.filter(
        company=request.user.company,
        quantity__lte=F('min_stock')
    ).count()
    
    recent_movements = Movement.objects.filter(
        company=request.user.company
    ).select_related('product', 'created_by').order_by('-created_at')[:10]
    
    stock_by_warehouse = Inventory.objects.filter(
        company=request.user.company
    ).values('warehouse__name').annotate(
        total=Sum('quantity')
    ).order_by('-total')[:5]
    
    # Productos con stock bajo
    low_stock_items = Inventory.objects.filter(
        company=request.user.company,
        quantity__lte=F('min_stock'),
        quantity__gt=0
    ).select_related('product', 'warehouse')[:10]
    
    context = {
        'total_products': total_products,
        'total_warehouses': total_warehouses,
        'low_stock_count': low_stock_count,
        'recent_movements': recent_movements,
        'stock_by_warehouse': stock_by_warehouse,
        'low_stock_items': low_stock_items,
    }
    return render(request, 'dashboard.html', context)

@login_required
@permission_required('reports.view_report', raise_exception=True)
def inventory_report_pdf(request):
    """Generar reporte de inventario en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    
    # Título
    title = Paragraph(f"Reporte de Inventario - {datetime.now().strftime('%d/%m/%Y %H:%M')}", title_style)
    elements.append(title)
    elements.append(Paragraph("<br/><br/>", styles['Normal']))
    
    # Datos
    inventories = Inventory.objects.filter(
        company=request.user.company,
        quantity__gt=0
    ).select_related('product', 'warehouse')
    
    data = [['Producto', 'SKU', 'Bodega', 'Cantidad', 'Stock Mínimo', 'Estado']]
    
    for inv in inventories:
        estado = "Bajo Stock" if inv.quantity <= inv.min_stock else "Normal"
        data.append([
            inv.product.name,
            inv.product.sku,
            inv.warehouse.name,
            str(inv.quantity),
            str(inv.min_stock),
            estado
        ])
    
    # Tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="inventory_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response

@login_required
@permission_required('reports.view_report', raise_exception=True)
def inventory_report_excel(request):
    """Generar reporte de inventario en Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Encabezados
    headers = ['Producto', 'SKU', 'Bodega', 'Cantidad', 'Stock Mínimo', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    inventories = Inventory.objects.filter(
        company=request.user.company,
        quantity__gt=0
    ).select_related('product', 'warehouse')
    
    for row, inv in enumerate(inventories, 2):
        estado = "Bajo Stock" if inv.quantity <= inv.min_stock else "Normal"
        
        ws.cell(row=row, column=1, value=inv.product.name)
        ws.cell(row=row, column=2, value=inv.product.sku)
        ws.cell(row=row, column=3, value=inv.warehouse.name)
        ws.cell(row=row, column=4, value=inv.quantity)
        ws.cell(row=row, column=5, value=inv.min_stock)
        
        cell = ws.cell(row=row, column=6, value=estado)
        if estado == "Bajo Stock":
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(color="9C0006")
        else:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell.font = Font(color="006100")
    
    # Ajustar ancho de columnas
    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inventory_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    
    return response

@login_required
@permission_required('reports.view_report', raise_exception=True)
def movements_report_pdf(request):
    """Generar reporte de movimientos en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    
    # Título
    title = Paragraph(f"Reporte de Movimientos - {datetime.now().strftime('%d/%m/%Y %H:%M')}", title_style)
    elements.append(title)
    elements.append(Paragraph("<br/><br/>", styles['Normal']))
    
    # Datos
    movements = Movement.objects.filter(
        company=request.user.company
    ).select_related('product', 'warehouse_from', 'warehouse_to', 'created_by').order_by('-created_at')[:100]
    
    data = [['Fecha', 'Tipo', 'Producto', 'Cantidad', 'Origen', 'Destino', 'Usuario']]
    
    for mov in movements:
        data.append([
            mov.created_at.strftime('%d/%m/%Y %H:%M'),
            mov.get_movement_type_display(),
            f"{mov.product.sku} - {mov.product.name}",
            str(mov.quantity),
            mov.warehouse_from.name if mov.warehouse_from else '-',
            mov.warehouse_to.name if mov.warehouse_to else '-',
            mov.created_by.username
        ])
    
    # Tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="movements_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response

@login_required
@permission_required('reports.view_report', raise_exception=True)
def movements_report_excel(request):
    """Generar reporte de movimientos en Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    # Encabezados
    headers = ['Fecha', 'Tipo', 'Producto', 'SKU', 'Cantidad', 'Origen', 'Destino', 'Usuario', 'Referencia']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    movements = Movement.objects.filter(
        company=request.user.company
    ).select_related('product', 'warehouse_from', 'warehouse_to', 'created_by').order_by('-created_at')[:1000]
    
    for row, mov in enumerate(movements, 2):
        ws.cell(row=row, column=1, value=mov.created_at.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=2, value=mov.get_movement_type_display())
        ws.cell(row=row, column=3, value=mov.product.name)
        ws.cell(row=row, column=4, value=mov.product.sku)
        ws.cell(row=row, column=5, value=mov.quantity)
        ws.cell(row=row, column=6, value=mov.warehouse_from.name if mov.warehouse_from else '-')
        ws.cell(row=row, column=7, value=mov.warehouse_to.name if mov.warehouse_to else '-')
        ws.cell(row=row, column=8, value=mov.created_by.username)
        ws.cell(row=row, column=9, value=mov.reference or '-')
    
    # Ajustar ancho de columnas
    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="movements_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    
    return response

@login_required
@permission_required('reports.view_report', raise_exception=True)
def kardex_report_pdf(request, product_id):
    """Generar reporte de Kardex por producto en PDF"""
    product = Product.objects.get(id=product_id, company=request.user.company)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    
    # Título
    title = Paragraph(f"Kardex - {product.name} ({product.sku}) - {datetime.now().strftime('%d/%m/%Y %H:%M')}", title_style)
    elements.append(title)
    elements.append(Paragraph("<br/><br/>", styles['Normal']))
    
    # Datos
    kardex_entries = Kardex.objects.filter(
        company=request.user.company,
        product=product
    ).select_related('warehouse', 'created_by').order_by('-created_at')[:100]
    
    data = [['Fecha', 'Tipo', 'Bodega', 'Entrada', 'Salida', 'Saldo', 'Usuario']]
    
    for entry in kardex_entries:
        data.append([
            entry.created_at.strftime('%d/%m/%Y %H:%M'),
            entry.get_movement_type_display(),
            entry.warehouse.name,
            str(entry.input_quantity) if entry.input_quantity > 0 else '-',
            str(entry.output_quantity) if entry.output_quantity > 0 else '-',
            str(entry.balance_quantity),
            entry.created_by.username
        ])
    
    # Tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="kardex_{product.sku}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response